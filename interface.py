import openpyxl

wb = openpyxl.load_workbook("Event_Manager.xlsx")
sheets = wb.sheetnames
sh1 = wb.active

def addit():
    m = int(input("Enter Month : "))
    d = int(input("Enter Date : "))
    eve = input("Enter Event : ")
    sh1.cell(row = d , column = m , value = eve)
    wb.save("Event_Manager.xlsx")
    print("Event added")


def checkit():
    c = int(input("Enter Month: "))
    r = int(input("Enter Date : "))
    print(sh1.cell(row=r, column=c).value)


from tkinter import *


window = Tk()

window.title('Event Manager')
window.geometry("300x200+10+20")

b=Button(window , text="Add entry" , height = 10 , bg = 'green', width = 10 , command = addit )
b.pack(side = LEFT)

c=Button(window , text="Check entry" , height = 10 , width = 10 , bg = 'green' , command = checkit)
c.pack(side = RIGHT)

window.mainloop()

